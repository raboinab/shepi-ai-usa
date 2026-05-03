"""
Generate static demo deliverables for the Shepi demo Export Center.

Outputs:
  public/demo/acme-sample-qoe.pdf       — watermarked sample QoE report
  public/demo/acme-sample-workbook.xlsx — watermarked sample workbook

Run: python3 scripts/generate-demo-deliverables.py
"""
import os
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = "public/demo"
os.makedirs(OUT_DIR, exist_ok=True)

# ---------------------------- PDF ----------------------------

COMPANY = "Acme Industrial Supply Co."
PERIOD = "Jan 2022 – Dec 2024 (TTM through Dec 2024)"

def watermark_canvas(canv, doc):
    canv.saveState()
    canv.setFont("Helvetica-Bold", 60)
    canv.setFillColorRGB(0.85, 0.85, 0.85, alpha=0.35)
    canv.translate(letter[0] / 2, letter[1] / 2)
    canv.rotate(35)
    canv.drawCentredString(0, 0, "DEMO — NOT FOR DISTRIBUTION")
    canv.rotate(-35)
    # Footer
    canv.setFont("Helvetica", 8)
    canv.setFillColorRGB(0.4, 0.4, 0.4, alpha=1)
    canv.drawString(-letter[0]/2 + 36, -letter[1]/2 + 24,
                    f"Shepi · Sample QoE Report · {COMPANY} · Demo preview")
    canv.drawRightString(letter[0]/2 - 36, -letter[1]/2 + 24,
                         f"Page {doc.page}")
    canv.restoreState()

def build_pdf():
    path = os.path.join(OUT_DIR, "acme-sample-qoe.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=22,
                        textColor=colors.HexColor("#0F172A"), spaceAfter=12)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=14,
                        textColor=colors.HexColor("#1E293B"), spaceAfter=8, spaceBefore=14)
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=10, leading=14,
                          textColor=colors.HexColor("#334155"))
    eyebrow = ParagraphStyle("Eyebrow", parent=styles["BodyText"], fontSize=9,
                             textColor=colors.HexColor("#64748B"),
                             spaceAfter=4, fontName="Helvetica-Bold")

    story = []

    # Cover
    story.append(Spacer(1, 1.5*inch))
    story.append(Paragraph("QUALITY OF EARNINGS REPORT", eyebrow))
    story.append(Paragraph(COMPANY, h1))
    story.append(Paragraph(PERIOD, body))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(
        "Prepared by Shepi · Analytical QoE software. "
        "This report presents adjustments and analytical findings. "
        "It is not an audit, review, or attestation engagement.",
        body))
    story.append(PageBreak())

    # Executive Summary
    story.append(Paragraph("Executive Summary", h2))
    story.append(Paragraph(
        f"{COMPANY} reported TTM revenue of $24.6M and reported EBITDA of $3.12M "
        "(12.7% margin). After analytical adjustments, Adjusted EBITDA is $3.84M "
        "(15.6% margin), an uplift of $720K driven primarily by owner compensation "
        "normalization, non-recurring legal expense, and personal expenses identified "
        "in the general ledger.", body))
    story.append(Spacer(1, 12))

    summary_data = [
        ["Metric", "Reported", "Adjusted", "Δ"],
        ["Revenue (TTM)", "$24,615,420", "$24,615,420", "—"],
        ["Gross Profit", "$8,743,210", "$8,743,210", "—"],
        ["Gross Margin", "35.5%", "35.5%", "—"],
        ["EBITDA", "$3,121,800", "$3,841,800", "+$720,000"],
        ["EBITDA Margin", "12.7%", "15.6%", "+290 bps"],
        ["Net Working Capital", "$2,847,300", "$2,847,300", "—"],
    ]
    t = Table(summary_data, colWidths=[2.2*inch, 1.4*inch, 1.4*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(PageBreak())

    # EBITDA Bridge
    story.append(Paragraph("EBITDA Adjustments Bridge", h2))
    story.append(Paragraph(
        "Each adjustment below was identified via Shepi's automated analysis of GL "
        "transactions, payroll records, and bank statements, then reviewed for support "
        "tier and re-occurrence likelihood.", body))
    story.append(Spacer(1, 10))

    bridge = [
        ["#", "Adjustment", "Class", "Amount", "Support"],
        ["1", "Owner compensation normalization", "Pro-forma", "+$420,000", "Tier 1"],
        ["2", "Non-recurring legal settlement (2023)", "Non-recurring", "+$185,000", "Tier 1"],
        ["3", "Personal expenses — owner credit card", "Personal/Discretionary", "+$78,400", "Tier 2"],
        ["4", "Related-party rent above market", "Related party", "+$54,200", "Tier 2"],
        ["5", "One-time consulting (CRM migration)", "Non-recurring", "+$38,500", "Tier 1"],
        ["6", "Inventory write-down (one-time)", "Non-recurring", "+$26,900", "Tier 2"],
        ["7", "Discretionary travel — owner family", "Personal/Discretionary", "+$19,800", "Tier 2"],
        ["8", "T&E — country club dues", "Personal/Discretionary", "+$14,400", "Tier 1"],
        ["", "Total adjustments", "", "+$720,000 (estimate)", ""],
    ]
    t2 = Table(bridge, colWidths=[0.3*inch, 2.6*inch, 1.4*inch, 1.2*inch, 0.7*inch])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("ALIGN", (4, 0), (4, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t2)
    story.append(PageBreak())

    # Narrative — Owner Comp
    story.append(Paragraph("Adjustment Detail · Owner Compensation Normalization", h2))
    story.append(Paragraph("RATIONALE", eyebrow))
    story.append(Paragraph(
        "The two majority owners are paid $620K combined in W-2 wages plus $180K in "
        "discretionary bonuses. Industry benchmarks for a CEO + COO at this revenue scale "
        "(Robert Half 2024, BLS OES) cluster at $360K–$400K combined. We normalize to "
        "$380K, yielding a $420K add-back. This adjustment is held at Tier 1 (highest "
        "support) — payroll register tied to W-2s, owner roles confirmed in org chart.",
        body))
    story.append(Spacer(1, 8))
    story.append(Paragraph("RISK CALLOUT", eyebrow))
    story.append(Paragraph(
        "A buyer may insist on a higher post-close compensation package to retain key "
        "staff currently absorbing owner workload. Diligence should confirm operational "
        "transition plan before fully crediting this adjustment.", body))
    story.append(PageBreak())

    # Risk indicators
    story.append(Paragraph("Risk Indicators", h2))
    risks = [
        ["Indicator", "Status", "Notes"],
        ["Customer concentration", "MEDIUM", "Top 3 customers = 41% of TTM revenue"],
        ["Cash proof of revenue", "PASS", "Bank deposits reconcile to GL within 0.3%"],
        ["GL anomalies (round-dollar JEs)", "LOW", "12 round-dollar entries, all <$5K, immaterial"],
        ["Related-party transactions", "MEDIUM", "$54K rent above market — see adj. #4"],
        ["Working capital trend", "PASS", "NWC stable, DSO 42 days (consistent)"],
        ["Inventory aging", "WATCH", "11% of inventory >180 days; potential reserve"],
    ]
    t3 = Table(risks, colWidths=[2.0*inch, 0.9*inch, 3.3*inch])
    t3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t3)
    story.append(Spacer(1, 0.4*inch))
    story.append(Paragraph("DISCLAIMER", eyebrow))
    story.append(Paragraph(
        "This is a sample report generated from synthetic data for product demonstration. "
        "Shepi provides analytical QoE software. We do not issue audit opinions, attestation "
        "reports, or formal CPA opinions through the platform.",
        body))

    doc.build(story, onFirstPage=watermark_canvas, onLaterPages=watermark_canvas)
    print(f"✓ {path}")

# ---------------------------- XLSX ----------------------------

THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="0F172A")
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
DEMO_FILL = PatternFill("solid", fgColor="FEF3C7")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")
        cell.border = BORDER

def style_body(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL

def build_xlsx():
    path = os.path.join(OUT_DIR, "acme-sample-workbook.xlsx")
    wb = Workbook()

    # DEMO sheet (first tab)
    demo = wb.active
    demo.title = "DEMO"
    demo["A1"] = "DEMO PREVIEW — NOT FOR DISTRIBUTION"
    demo["A1"].font = Font(bold=True, size=18, color="92400E")
    demo["A1"].fill = DEMO_FILL
    demo.merge_cells("A1:F1")
    demo.row_dimensions[1].height = 36
    demo["A1"].alignment = Alignment(horizontal="center", vertical="center")

    notes = [
        "",
        f"Company: {COMPANY}",
        f"Period:  {PERIOD}",
        "",
        "This workbook is a sample produced from synthetic data.",
        "Sign up at shepi.ai to generate your own QoE workbook from real client data.",
        "",
        "Sheets included in this preview:",
        "  • Income Statement",
        "  • EBITDA Bridge",
        "  • Adjustments Detail",
    ]
    for i, line in enumerate(notes, start=2):
        demo[f"A{i}"] = line
        demo[f"A{i}"].font = Font(size=11, color="334155")
    demo.column_dimensions["A"].width = 80

    # Income Statement
    is_ws = wb.create_sheet("DEMO · Income Statement")
    is_data = [
        ["", "FY2022", "FY2023", "FY2024 (TTM)"],
        ["Revenue", 21_840_120, 23_215_400, 24_615_420],
        ["Cost of Goods Sold", -14_120_400, -14_980_220, -15_872_210],
        ["Gross Profit", "=B2+B3", "=C2+C3", "=D2+D3"],
        ["Gross Margin %", "=B4/B2", "=C4/C2", "=D4/D2"],
        ["", "", "", ""],
        ["Operating Expenses", "", "", ""],
        ["  Salaries & Wages", -3_840_000, -4_120_000, -4_385_000],
        ["  Owner Compensation", -780_000, -795_000, -800_000],
        ["  Rent", -420_000, -432_000, -445_000],
        ["  Professional Fees", -185_000, -310_000, -142_000],
        ["  Travel & Entertainment", -98_000, -112_000, -134_200],
        ["  Other OpEx", -612_000, -638_000, -677_220],
        ["Total OpEx", "=SUM(B8:B13)", "=SUM(C8:C13)", "=SUM(D8:D13)"],
        ["", "", "", ""],
        ["Operating Income", "=B4+B14", "=C4+C14", "=D4+D14"],
        ["Depreciation & Amortization", -185_000, -195_000, -210_000],
        ["EBITDA", "=B16-B17", "=C16-C17", "=D16-D17"],
        ["EBITDA Margin %", "=B18/B2", "=C18/C2", "=D18/D2"],
    ]
    for r, row in enumerate(is_data, start=1):
        for c, val in enumerate(row, start=1):
            is_ws.cell(row=r, column=c, value=val)
    style_header(is_ws, 1, 4)
    style_body(is_ws, 2, len(is_data), 4)
    # Number formats
    for r in [2, 3, 4, 8, 9, 10, 11, 12, 13, 14, 16, 17, 18]:
        for c in [2, 3, 4]:
            is_ws.cell(row=r, column=c).number_format = '$#,##0;($#,##0);-'
    for r in [5, 19]:
        for c in [2, 3, 4]:
            is_ws.cell(row=r, column=c).number_format = '0.0%'
    is_ws.column_dimensions["A"].width = 32
    for col in ["B", "C", "D"]:
        is_ws.column_dimensions[col].width = 18

    # EBITDA Bridge
    br = wb.create_sheet("DEMO · EBITDA Bridge")
    br_data = [
        ["EBITDA Bridge — TTM Dec 2024", "", ""],
        ["", "Amount", "Cumulative"],
        ["Reported EBITDA", 3_121_800, "=B3"],
        ["+ Owner compensation normalization", 420_000, "=C3+B4"],
        ["+ Non-recurring legal settlement", 185_000, "=C4+B5"],
        ["+ Personal expenses (owner CC)", 78_400, "=C5+B6"],
        ["+ Related-party rent above market", 54_200, "=C6+B7"],
        ["+ One-time CRM consulting", 38_500, "=C7+B8"],
        ["+ Inventory write-down (one-time)", 26_900, "=C8+B9"],
        ["+ Discretionary owner travel", 19_800, "=C9+B10"],
        ["+ T&E — country club dues", 14_400, "=C10+B11"],
        ["Adjusted EBITDA", "=SUM(B3:B11)", "=C11"],
    ]
    for r, row in enumerate(br_data, start=1):
        for c, val in enumerate(row, start=1):
            br.cell(row=r, column=c, value=val)
    br["A1"].font = Font(bold=True, size=14, color="0F172A")
    br.merge_cells("A1:C1")
    style_header(br, 2, 3)
    style_body(br, 3, len(br_data), 3)
    for r in range(3, 13):
        for c in [2, 3]:
            br.cell(row=r, column=c).number_format = '$#,##0;($#,##0);-'
    br.cell(row=12, column=1).font = Font(bold=True, size=11)
    br.cell(row=12, column=2).font = Font(bold=True, size=11)
    br.column_dimensions["A"].width = 42
    br.column_dimensions["B"].width = 18
    br.column_dimensions["C"].width = 18

    # Adjustments Detail
    adj = wb.create_sheet("DEMO · Adjustments")
    adj_data = [
        ["#", "Adjustment", "Class", "Amount", "Support Tier", "Rationale"],
        [1, "Owner compensation normalization", "Pro-forma", 420_000, "Tier 1",
         "Combined W-2 + bonus $800K vs. industry benchmark $380K (BLS OES)."],
        [2, "Non-recurring legal settlement", "Non-recurring", 185_000, "Tier 1",
         "2023 wrongful termination settlement; one-time, no future recurrence."],
        [3, "Personal expenses (owner CC)", "Personal/Discretionary", 78_400, "Tier 2",
         "Identified via GL anomaly detection; meals, retail, family travel."],
        [4, "Related-party rent above market", "Related party", 54_200, "Tier 2",
         "Property owned by majority shareholder; rent 18% above market comps."],
        [5, "One-time CRM consulting", "Non-recurring", 38_500, "Tier 1",
         "2023 Salesforce migration; vendor invoice confirms one-time scope."],
        [6, "Inventory write-down", "Non-recurring", 26_900, "Tier 2",
         "Q3 2023 obsolete SKU reserve; not recurring per management."],
        [7, "Discretionary owner travel", "Personal/Discretionary", 19_800, "Tier 2",
         "Family vacation booked through corporate travel account."],
        [8, "T&E — country club dues", "Personal/Discretionary", 14_400, "Tier 1",
         "Annual dues; classified as discretionary owner benefit."],
        ["", "Total", "", "=SUM(D2:D9)", "", ""],
    ]
    for r, row in enumerate(adj_data, start=1):
        for c, val in enumerate(row, start=1):
            adj.cell(row=r, column=c, value=val)
    style_header(adj, 1, 6)
    style_body(adj, 2, len(adj_data), 6)
    for r in range(2, 11):
        adj.cell(row=r, column=4).number_format = '$#,##0;($#,##0);-'
    adj.cell(row=10, column=1).font = Font(bold=True)
    adj.cell(row=10, column=2).font = Font(bold=True)
    adj.cell(row=10, column=4).font = Font(bold=True)
    adj.column_dimensions["A"].width = 5
    adj.column_dimensions["B"].width = 36
    adj.column_dimensions["C"].width = 22
    adj.column_dimensions["D"].width = 14
    adj.column_dimensions["E"].width = 12
    adj.column_dimensions["F"].width = 60
    for r in range(2, 10):
        adj.row_dimensions[r].height = 32
        adj.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(path)
    print(f"✓ {path}")

if __name__ == "__main__":
    build_pdf()
    build_xlsx()
